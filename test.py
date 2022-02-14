from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils.dataframe import dataframe_to_rows
wb2 = load_workbook('13F-data.xlsx')


ws = wb2["Stocks By Quarters"]

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

if (ws['F6'].value < 0):
    ws['F6'] = -ws['F6'].value
    ws['A1'].fill = redFill

ws['U14'] = "=SUM(B2:E2)%"
print(ws['U14'].value)

ws['A1'].font = Font(color = "FF0000")
