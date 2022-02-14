from ntpath import join
import pandas as pd
import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

sec_url = 'https://www.sec.gov'

MERGE_CONFIG = {
    'Stocks By Quarters': 'A1:K1',
    'Bought Stocks': 'A1:H1',
    'Sold Stocks': 'A1:H1',
}

def difference_color_function(ws, len):
    def change_cell_color(cell_id):
        red = '00FF0000'
        green = '0000FF00'
        try:    
            if ws[cell_id].value < 0:
                ws[cell_id] = round(-ws[cell_id].value * 100, 2)
                ws[cell_id].font = Font(color = "00FF0000") 
                return            
            ws[cell_id] = round(ws[cell_id].value * 100, 2)
            ws[cell_id].font = Font(color = "0000FF00")             
        except Exception as e:
            print(e)
    
    dif_cell = ['G', 'J']
    for i in range(4, 4+len+1):
        for cell_name in dif_cell:
            change_cell_color(f"{cell_name}{i}")

COLOR_CONFIG = {
    'Stocks By Quarters': difference_color_function,
    'Bought Stocks': lambda x, y: x,
    'Sold Stocks': lambda x, y: x,
}

def get_request(url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36',
        'Accept-Encoding': 'gzip, deflate, br',
        'HOST': 'www.sec.gov',
    }
    return requests.get(url, headers=headers)

def create_url(cik):
    return 'https://www.sec.gov/cgi-bin/browse-edgar?CIK={}&owner=exclude&action=getcompany&type=13F-HR'.format(cik)

def get_user_input():
    cik = input("Enter 10-digit CIK number: ")
    return cik

def scrap_company_report(requested_cik):
    # Find mutual fund by CIK number on EDGAR
    response = get_request(create_url(requested_cik))
    soup = BeautifulSoup(response.text, "html.parser")
    tags = soup.findAll('a', id="documentsbutton")
    last_report = (sec_url + tags[0]['href'])
    previous_report = (sec_url + tags[1]['href'])
    last = scrap_report_by_url(last_report, "last_report")
    previous = scrap_report_by_url(previous_report, "previous_report")
    return last, previous

def scrap_report_by_url(url, filename):
    response_two = get_request(url)
    soup_two = BeautifulSoup(response_two.text, "html.parser")
    tags_two = soup_two.findAll('a', attrs={'href': re.compile('xml')})
    xml_url = tags_two[3].get('href')
    response_xml = get_request(sec_url + xml_url)
    soup_xml = BeautifulSoup(response_xml.content, "lxml")
    return fetch_xml_data(soup_xml)

def fetch_xml_data(soup_xml):
    columns = [
        "Name of Issuer",
        "CUSIP",
        "Value",
        "Shares",
        "Investment Discretion",
    ]
    issuers = soup_xml.body.findAll(re.compile('nameofissuer'))
    cusips = soup_xml.body.findAll(re.compile('cusip'))
    values = soup_xml.body.findAll(re.compile('value'))
    sshprnamts = soup_xml.body.findAll('sshprnamt')
    investmentdiscretions = soup_xml.body.findAll(re.compile('investmentdiscretion'))
    df = pd.DataFrame(columns= columns)
    for issuer, cusip, value, sshprnamt, investmentdiscretion, in zip(issuers, cusips, values, sshprnamts, investmentdiscretions):
        row_df = pd.DataFrame({
            "Name of Issuer": [issuer.text],
            "CUSIP": [cusip.text],
            "Value": [int(value.text)],
            "Shares": [int(sshprnamt.text)],
            "Investment Discretion": [investmentdiscretion.text],
        })
        df = pd.concat([df, row_df], ignore_index = True, axis = 0)

    df["Percentage Of holdings"] = df.Value / df.Value.sum() 
    return df

def get_new_stocks(df1, df2, sold=False):
    '''
        this method will find stocks from df1, which does not occur in df2
    '''
    fix_dict = {
    }
    for _, row in df2.iterrows():
        cusip = row["CUSIP"]
        investment_type = row["Investment Discretion"]
        id = f"{cusip}-{investment_type}"
        fix_dict[id] = True

    new_comers_df = pd.DataFrame(columns=[*df1.columns,"type"])

    for _, row in df1.iterrows():
        cusip = row["CUSIP"]
        investment_type = row["Investment Discretion"]
        id = f"{cusip}-{investment_type}"
        if id not in fix_dict:
            new_comers_df = new_comers_df.append(row)

    new_comers_df["type"] = "sold" if sold else "bought"
    return new_comers_df

def join_stocks(last_quarter, previous_quarter):
    '''
        this method will join stocks from df1, which does not occur in df2
    '''
    column_ordering = [
        'Name of Issuer', 'CUSIP', 'Investment Discretion', 
        'Shares Q1', 'Shares Q2', 'Shares Difference %',
        'Value Q1', 'Value Q2', "Value Difference %",
        'Percentage Of holdings Q2', 'Percentage Of holdings Q1',
    ]

    # Merge the dataframes into another dataframe based on PERSONID and Badge_ID
    not_rename_columns = ["CUSIP", "Investment Discretion", "Name of Issuer"]
    last_quarter.rename(index=str, columns=dict([(col, col+' Q1') for col in last_quarter if col not in not_rename_columns]), inplace=True)
    previous_quarter.rename(index=str, columns=dict([(col, col+' Q2') for col in previous_quarter if col not in not_rename_columns]), inplace=True)
    join_df = pd.merge(last_quarter, previous_quarter, how='inner', on=['Name of Issuer', 'CUSIP', 'Investment Discretion'])
    join_df["Shares Difference %"] = (join_df["Shares Q1"] - join_df["Shares Q2"]) / join_df["Shares Q2"]
    join_df["Value Difference %"] = (join_df["Value Q1"] - join_df["Value Q2"]) / join_df["Value Q2"]

    join_df = join_df.reindex(columns= column_ordering)
    return join_df


def make_excel(last, previous):
    wb = Workbook()
    ws0 = wb.create_sheet("Stocks By Quarters")
    join = join_stocks(last, previous)
    df_to_excel(join, ws0, "Stocks By Quarters")

    ws1 = wb.create_sheet("Bought Stocks")
    boughts = get_new_stocks(last, previous)
    df_to_excel(boughts, ws1, "Bought Stocks")
    
    ws2 = wb.create_sheet("Sold Stocks")
    solds = get_new_stocks(previous, last, True)
    df_to_excel(solds, ws2, "Sold Stocks")

    del wb["Sheet"]    
    wb.save(filename = '13F-data.xlsx')

def df_to_excel(df, ws, name):
    ws.merge_cells(MERGE_CONFIG[name])
    ws['A1'] = name
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'

    COLOR_CONFIG[name](ws, df.shape[0])

requested_cik = get_user_input()
last, previous =  scrap_company_report(requested_cik)

make_excel(last, previous)